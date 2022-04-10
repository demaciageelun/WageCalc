import json

import openpyxl
from django import forms
from django.contrib import admin
from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.contrib.auth.forms import ReadOnlyPasswordHashField
from django.core.exceptions import ValidationError
from django.db.models import F
from django.http import JsonResponse, FileResponse
from django.utils.html import format_html
from simpleui.admin import AjaxAdmin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin

import wage
from .models import *
import datetime as firstdate
from .functions import monthcalc, verifyIdName

period_list = WagePeriod.objects.values('id', 'period_name')
global options_list
options_list = []
for data in period_list:
    temp = {'key': data['id'], 'label': data['period_name']}
    options_list.append(temp)


def get_job_status(num):
    numbers = {
        1: "在职",
        2: "离职",
        99: "黑名单",
    }
    return numbers.get(num, None)


# Register your models here.
# 周期表
class WagePeriodAdmin(AjaxAdmin):
    list_display = ['period_name']
    search_fields = ['period_name']


# 部门表
class WageDeptmentAdmin(AjaxAdmin):
    list_display = ['dept_id', 'dept_name', 'dept_short_name', 'dept_base']
    search_fields = ['dept_id', 'dept_name', 'dept_base', 'dept_short_name']
    list_filter = ['dept_base', 'dept_short_name']


# 岗位表
class WagePositionAdmin(AjaxAdmin):
    list_display = ['posi_id', 'posi_name']
    search_fields = ['posi_id', 'posi_name']
    # list_filter = ['posi_id', 'posi_name']


# 员工表
class WageEmployeeAdmin(AjaxAdmin):
    list_display = ['emp_id', 'emp_name', 'emp_dept', 'emp_posi', 'emp_rank', 'glgs', 'emp_entry_date', 'emp_in_date',
                    'emp_job_type', 'emp_job_status', 'emp_leave_date', 'emp_leave_bl_date', 'emp_job_level', 'emp_dlidl', 'emp_pay_type', 'emp_base']
    search_fields = ['emp_id', 'emp_name']
    list_filter = ['emp_dept__dept_short_name', 'emp_posi__posi_name', 'emp_job_type', 'emp_job_status', 'emp_rank',
                   'emp_dept__dept_base', 'emp_job_level', 'emp_dlidl', 'emp_pay_type']
    actions = ['download']

    def glgs(self, obj):
        return obj.emp_dept.dept_base

    glgs.short_description = '管理归属'

    def get_actions(self, request):
        actions = super(WageEmployeeAdmin, self).get_actions(request)
        if not request.user.has_perm('wage.can_download_employee'):
            del actions['download']
        return actions

    # 重写显示的数据内容，可根据不同登录用户，显示不同基地的人员信息。
    def get_queryset(self, request):
        qs = super(WageEmployeeAdmin, self).get_queryset(request)
        if not request.user.is_superuser:
            myUser = MyUser.objects.get(user_id=request.user.id)
            deptIdList = myUser.department.all().values_list('id')
            rankId = myUser.rank.all().values_list('id')
            return qs.filter(emp_dept_id__in=deptIdList, emp_rank_id__in=rankId)
        else:
            return qs

    def download(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        row1 = ['工号', '姓名', '部门', '岗位', '集团入职日期', '入职日期', '离职结薪日期', '离职办理日期', '在职状态', '在职类型', '合同归属', '管理归属']
        ws.append(row1)
        for datas in queryset:
            row2 = [
                str(datas.emp_id) if str(datas.emp_id) != "None" else "",  # 工号
                str(datas.emp_name) if str(datas.emp_name) != "None" else "",  # 姓名
                str(datas.emp_dept) if str(datas.emp_dept) != "None" else "",  # 部门
                str(datas.emp_posi) if str(datas.emp_posi) != "None" else "",  # 岗位
                str(datas.emp_entry_date) if str(datas.emp_entry_date) != "None" else "",  # 集团入职日期
                str(datas.emp_in_date) if str(datas.emp_in_date) != "None" else "",  # 入职日期
                str(datas.emp_leave_date) if str(datas.emp_leave_date) != "None" else "",  # 离职结薪日期
                str(datas.emp_leave_bl_date) if str(datas.emp_leave_bl_date) != "None" else "",  # 离职办理日期
                get_job_status(datas.emp_job_status),  # 在职状态
                str(datas.emp_job_type) if str(datas.emp_job_type) != "None" else "",  # 在职类型
                str(datas.emp_rank) if str(datas.emp_rank) != "None" else "",  # 合同归属
                str(datas.emp_dept.dept_base) if str(datas.emp_dept.dept_base) != "None" else "",  # 管理归属
            ]
            ws.append(row2)
        wb.save('static/download/emp.xlsx')
        #     下载
        file = open('static/download/emp.xlsx', 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="emp.xlsx"'
        return response

        # 显示的文本，与django admin一致

    download.short_description = '下载数据'
    # icon，参考element-ui icon与https://fontawesome.com
    # custom_button.icon = 'fas fa-audio-description'
    # 指定element-ui的按钮类型，参考https://element.eleme.cn/#/zh-CN/component/button
    download.type = 'success'


# 费用项目表
class WageExpenseAdmin(AjaxAdmin):
    list_display = ['expense_name']
    search_fields = ['expense_name']


# 费用项目引入表
class WageExpenseIntoAdmin(AjaxAdmin):
    list_display = ['into_emp', 'into_emp_ids', 'dept', 'position', 'into_periods', 'into_expense', 'into_reason',
                    'into_calc_elem_1', 'into_calc_elem_2', 'into_calc_elem_other', 'into_money', 'into_effecte_date', 'into_expir_date']
    list_display_links = ['into_emp', 'into_emp_ids', 'dept', 'position', 'into_periods', 'into_expense', 'into_reason',
                          'into_calc_elem_1', 'into_calc_elem_2', 'into_calc_elem_other', 'into_money', 'into_effecte_date', 'into_expir_date']
    actions = ['upload_file', 'download_templates', 'download']
    list_filter = ['into_emp__emp_dept__dept_short_name', 'into_emp__emp_posi', 'into_periods', 'into_expense']
    raw_id_fields = ['into_emp', 'into_expense']
    search_fields = ['into_emp__emp_id', 'into_emp__emp_name']

    def get_queryset(self, request):
        qs = super(WageExpenseIntoAdmin, self).get_queryset(request)
        # 是否为管理员，否需要过滤部门基地；是返回全部数据
        if not request.user.is_superuser:
            myUser = MyUser.objects.get(user_id=request.user.id)
            deptIdList = myUser.department.all().values_list('id')
            rankId = myUser.rank.all().values_list('id')
            # 是否可以查看所有信息，包含别人上传的
            if request.user.myuser.is_full is False:
                return qs.filter(into_emp__emp_dept_id__in=deptIdList, into_emp__emp_rank_id__in=rankId, into_creater=str(request.user))
            else:
                return qs.filter(into_emp__emp_dept_id__in=deptIdList, into_emp__emp_rank_id__in=rankId)
        else:
            return qs

    def get_readonly_fields(self, request, obj=None):
        if request.user.is_superuser:
            return 'into_creater', 'into_create_time'
        else:
            return 'into_emp', 'dept', 'position', 'into_periods', 'into_expense', 'into_reason', 'into_calc_elem_1', 'into_calc_elem_2', 'into_calc_elem_other', 'into_money', 'into_effecte_date', 'into_expir_date', 'fix_creater', 'fix_create_time'

    def into_emp_ids(self, obj):
        return obj.into_emp.emp_id

    into_emp_ids.short_description = '工号'

    def dept(self, obj):
        return obj.into_emp.emp_dept.dept_short_name

    dept.short_description = '部门'

    def position(self, obj):
        return obj.into_emp.emp_posi

    position.short_description = '岗位'

    # 根据登录账号判断显示的按钮
    def get_actions(self, request):
        actions = super(WageExpenseIntoAdmin, self).get_actions(request)
        if not request.user.has_perm('wage.can_import_expense'):
            del actions['upload_file']
        return actions

    def download_templates(self, request, queryset):
        file = open('static/templates/expense.xlsx', 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="expense.xlsx"'
        return response

    download_templates.short_description = '下载模板'
    download_templates.type = 'warning'
    download_templates.icon = 'el-icon-paperclip'
    download_templates.acts_on_all = True

    def upload_file(self, request, queryset):
        # 这里的upload 就是和params中配置的key一样
        upload = request.FILES['upload']
        # 将上传的文件保存到static中
        now_time = str(firstdate.datetime.now().strftime("%Y-%m-%d_%H.%M.%S"))
        WageLog.objects.create(log_user=str(request.user), log_model='费用引入', log_path='static/expense/', log_file_name=now_time + str(request.user) + upload.name)
        with open('static/expense/' + now_time + str(request.user) + upload.name, 'wb') as f:
            for line in upload.chunks():
                f.write(line)
        f.close()
        workbook = openpyxl.load_workbook('static/expense/' + now_time + str(request.user) + upload.name)
        # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
        table = workbook.active
        rows = table.max_row
        cols = table.max_column
        # 获取标题栏数据
        head_data = []
        for datas in range(cols):
            head_data.append(table.cell(1, datas + 1).value)
        if '工号' in head_data and '姓名' in head_data and '周期' in head_data and '成本归属' in head_data and '公司' in head_data and '项目' in head_data and '原因' in head_data and '计算要素1' in head_data and '计算要素2' in head_data and '其他计算要素' in head_data and '金额' in head_data and '生效日期' in head_data and '失效日期' in head_data:
            # 创建一个字符串来手机所有错误数据
            error_str = ''
            for datas in range(rows):
                if datas > 0:
                    emp_id = table.cell(datas + 1, head_data.index('工号') + 1).value
                    if type(emp_id) == int:
                        emp_name = table.cell(datas + 1, head_data.index('姓名') + 1).value
                        sql_name = WageEmployee.objects.raw("select id,emp_name from wage_employee where emp_id = '" + str(emp_id) + "'")
                        if len(sql_name) > 0 and str(sql_name[0]) == emp_name:
                            expense_name = table.cell(datas + 1, head_data.index('项目') + 1).value
                            period_name = table.cell(datas + 1, head_data.index('周期') + 1).value
                            into_emp_id = WageEmployee.objects.filter(emp_id=int(emp_id)).values('id')[0]['id']
                            into_expense_id = WageExpense.objects.filter(expense_name=expense_name).values('id')
                            if len(into_expense_id) > 0:
                                into_expense_id = into_expense_id[0]['id']
                                WageExpenseInto.objects.update_or_create(
                                    defaults={
                                        'into_emp_id': into_emp_id,
                                        'into_periods': period_name,
                                        'into_expense_id': into_expense_id,
                                        'into_reason': table.cell(datas + 1, head_data.index('原因') + 1).value,
                                        'into_calc_elem_1': table.cell(datas + 1, head_data.index('计算要素1') + 1).value,
                                        'into_calc_elem_2': table.cell(datas + 1, head_data.index('计算要素2') + 1).value,
                                        'into_calc_elem_other': table.cell(datas + 1, head_data.index('其他计算要素') + 1).value,
                                        'into_money': table.cell(datas + 1, head_data.index('金额') + 1).value,
                                        'into_effecte_date': table.cell(datas + 1, head_data.index('生效日期') + 1).value,
                                        'into_expir_date': table.cell(datas + 1, head_data.index('失效日期') + 1).value,
                                        'into_creater': str(request.user),
                                        'into_create_time': firstdate.datetime.now()

                                    },
                                    into_emp_id=into_emp_id,
                                    into_periods=period_name,
                                    into_money=table.cell(datas + 1, head_data.index('金额') + 1).value,
                                    into_expense_id=into_expense_id,
                                    into_reason=table.cell(datas + 1, head_data.index('原因') + 1).value,
                                    into_calc_elem_1=table.cell(datas + 1, head_data.index('计算要素1') + 1).value,
                                    into_calc_elem_2=table.cell(datas + 1, head_data.index('计算要素2') + 1).value,
                                    into_calc_elem_other=table.cell(datas + 1, head_data.index('其他计算要素') + 1).value,
                                )
                            else:
                                error_str += "第" + str(datas) + "行费用项目不存在，请修正。"
                                continue
                        else:
                            error_str += "第" + str(datas) + "行工号和姓名不匹配，请修正。"
                            continue
                    else:
                        error_str += "第" + str(datas) + "行有空字符串，请整行删除。"
                        continue
            # 如果错误
            if error_str == "":
                status = 'success'
                msg = '上传成功！'
            else:
                status = 'error'
                msg = error_str
            return JsonResponse(data={
                'status': status,
                'msg': str(msg)
            })
        else:
            status = 'error'
            msg = '列名错误，请使用下载的文档进行导入！'
            return JsonResponse(data={
                'status': status,
                'msg': str(msg)
            })

    upload_file.short_description = '文件上传'
    upload_file.type = 'success'
    upload_file.icon = 'el-icon-upload'
    upload_file.enable = True

    upload_file.layer = {
        'params': [{
            'type': 'file',
            'key': 'upload',
            'label': '文件'
        }]
    }

    def download(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        row1 = ['工号', '姓名', '周期', '成本归属', '员工状态', '在职状态', '入职日期', '离职结薪日期', '公司', '部门简称', '岗位名称', '项目', '原因', '计算要素1', '计算要素2', '其他计算要素', '金额', '生效日期', '失效日期']
        ws.append(row1)
        for datas in queryset:
            print(str(datas.into_emp.emp_leave_date))
            row2 = [
                int(datas.into_emp.emp_id),  # 工号
                str(datas.into_emp.emp_name),  # 姓名
                str(datas.into_periods),  # 周期
                str(datas.into_periods),  # 成本归属
                str(get_job_status(datas.into_emp.emp_job_status)),  # 员工状态
                str(datas.into_emp.emp_job_type),  # 在职状态
                str(datas.into_emp.emp_entry_date),  # 入职日期
                str(datas.into_emp.emp_leave_date) if str(datas.into_emp.emp_leave_date) != "None" else "",  # 离职结薪日期
                str(datas.into_emp.emp_rank),  # 公司
                str(datas.into_emp.emp_dept.dept_short_name),  # 部门简称
                str(datas.into_emp.emp_posi),  # 岗位名称
                str(datas.into_expense),  # 项目
                str(datas.into_reason) if str(datas.into_reason) != "None" else "",  # 原因
                str(datas.into_calc_elem_1) if str(datas.into_calc_elem_1) != "None" else "",  # 计算要素1
                str(datas.into_calc_elem_2) if str(datas.into_calc_elem_2) != "None" else "",  # 计算要素2
                str(datas.into_calc_elem_other) if str(datas.into_calc_elem_other) != "None" else "",  # 其他计算要素
                float(datas.into_money),  # 金额
                str(datas.into_effecte_date) if str(datas.into_effecte_date) != "None" else "",  # 生效日期
                str(datas.into_expir_date) if str(datas.into_expir_date) != "None" else "",  # 失效日期

            ]
            ws.append(row2)
        wb.save('static/download/expenseInto.xlsx')
        #     下载
        file = open('static/download/expenseInto.xlsx', 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="expenseInto.xlsx"'
        return response

    download.short_description = '下载数据'
    download.type = 'success'


# 绩效引入表
class WagePerformanceIntoAdmin(AjaxAdmin):
    list_display = ['perfor_emp', 'perfor_emp_ids', 'dept', 'perfor_center', 'position', 'perfor_periods', 'perfor_period_type',
                    'perfor_costs_attach', 'perfor_proof_ip', 'perfor_proof_dp', 'perfor_proof_cp', 'perfor_score_ip', 'perfor_score_dp', 'perfor_score_cp',
                    'perfor_ratio_ip', 'perfor_ratio_dp', 'perfor_ratio_cp', 'perfor_ratio_tp', 'perfor_result']
    list_display_links = ['perfor_emp', 'perfor_emp_ids', 'dept', 'perfor_center', 'position', 'perfor_periods', 'perfor_period_type',
                          'perfor_costs_attach', 'perfor_proof_ip', 'perfor_proof_dp', 'perfor_proof_cp', 'perfor_score_ip', 'perfor_score_dp', 'perfor_score_cp',
                          'perfor_ratio_ip', 'perfor_ratio_dp', 'perfor_ratio_cp', 'perfor_ratio_tp', 'perfor_result']
    raw_id_fields = ['perfor_emp']
    list_filter = ['perfor_emp__emp_dept__dept_short_name', 'perfor_emp__emp_posi', 'perfor_periods', 'perfor_costs_attach', 'perfor_center']
    actions = ['upload_file', 'download_templates', 'download']
    search_fields = ['perfor_emp__emp_id', 'perfor_emp__emp_name']

    def get_readonly_fields(self, request, obj=None):
        return 'perfor_creater', 'perfor_create_time'

    def get_queryset(self, request):
        qs = super(WagePerformanceIntoAdmin, self).get_queryset(request)
        if not request.user.is_superuser:
            myUser = MyUser.objects.get(user_id=request.user.id)
            deptIdList = myUser.department.all().values_list('id')
            rankId = myUser.rank.all().values_list('id')
            if request.user.myuser.is_full is False:
                return qs.filter(perfor_emp__emp_dept_id__in=deptIdList, perfor_emp__emp_rank_id__in=rankId, perfor_creater=str(request.user))
            else:
                return qs.filter(perfor_emp__emp_dept_id__in=deptIdList, perfor_emp__emp_rank_id__in=rankId)
        else:
            return qs

    # 根据登录账号判断显示的按钮
    def get_actions(self, request):
        actions = super(WagePerformanceIntoAdmin, self).get_actions(request)
        if not request.user.has_perm('wage.can_import_performance'):
            del actions['upload_file']
        return actions

    def get_readonly_fields(self, request, obj=None):
        if request.user.is_superuser:
            return 'perfor_creater', 'perfor_create_time'
        else:
            return 'perfor_emp', 'perfor_center', 'perfor_periods', 'perfor_period_type', 'perfor_costs_attach', 'perfor_proof_ip', 'perfor_proof_dp', 'perfor_proof_cp', 'perfor_score_ip', 'perfor_score_dp', 'perfor_score_cp', 'perfor_ratio_ip', 'perfor_ratio_dp', 'perfor_ratio_cp', 'perfor_ratio_tp', 'perfor_result', 'perfor_creater', 'perfor_create_time'

    def perfor_emp_ids(self, obj):
        return obj.perfor_emp.emp_id

    perfor_emp_ids.short_description = '工号'

    def dept(self, obj):
        return obj.perfor_emp.emp_dept.dept_short_name

    dept.short_description = '部门'

    def position(self, obj):
        return obj.perfor_emp.emp_posi

    position.short_description = '岗位'

    def download_templates(self, request, queryset):
        file = open('static/templates/performance.xlsx', 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="performance.xlsx"'
        return response

    download_templates.short_description = '下载模板'
    download_templates.type = 'warning'
    download_templates.icon = 'el-icon-paperclip'
    download_templates.acts_on_all = True

    def upload_file(self, request, queryset):
        # 这里的upload 就是和params中配置的key一样
        upload = request.FILES['upload']
        now_time = str(firstdate.datetime.now().strftime("%Y-%m-%d_%H.%M.%S"))
        WageLog.objects.create(log_user=str(request.user), log_model='绩效引入', log_path='static/performance/', log_file_name=now_time + str(request.user) + upload.name)
        # 将上传的文件保存到static中
        with open('static/performance/' + now_time + str(request.user) + upload.name, 'wb') as f:
            for line in upload.chunks():
                f.write(line)
        f.close()
        workbook = openpyxl.load_workbook('static/performance/' + now_time + str(request.user) + upload.name)
        # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
        table = workbook.active
        rows = table.max_row
        cols = table.max_column
        # 获取标题栏数据
        head_data = []
        for datas in range(cols):
            head_data.append(table.cell(1, datas + 1).value)
        # 首先判断列名是否都正确，正确则进入循环，错误就返回报错
        if '工号' in head_data and '姓名' in head_data and '中心别' in head_data and '周期名称' in head_data and '周期类型' in head_data and '成本归属' in head_data and '公司' in head_data and '个人绩效占比' in head_data and '部门绩效占比' in head_data and '公司绩效占比' in head_data and '个人绩效分数' in head_data and '部门绩效分数' in head_data and '公司绩效分数' in head_data and '个人绩效系数' in head_data and '部门绩效系数' in head_data and '公司绩效系数' in head_data:
            # 创建一个字符串来手机所有错误数据
            error_str = ''
            for datas in range(rows):
                if datas > 0:
                    emp_id = table.cell(datas + 1, head_data.index('工号') + 1).value
                    if type(emp_id) == int:
                        perfor_emp_id = WageEmployee.objects.filter(emp_id=int(emp_id)).values('id')[0]['id']
                        sql_name = WageEmployee.objects.raw("select id,emp_name from wage_employee where emp_id = '" + str(emp_id) + "'")
                        perfor_name = table.cell(datas + 1, head_data.index('姓名') + 1).value
                        if len(sql_name) > 0 and str(sql_name[0]) == perfor_name:
                            perfor_center = table.cell(datas + 1, head_data.index('中心别') + 1).value if table.cell(datas + 1, head_data.index('中心别') + 1).value is not None else ""
                            period_name = table.cell(datas + 1, head_data.index('周期名称') + 1).value
                            # 判断是否为日期格式（datetime），不是的话报错，是的话继续循环
                            if type(period_name) == int:
                                error_str += "第" + str(datas) + "行日期格式错误，请以文本格式，不要使用自定义格式。"
                                continue
                            else:
                                perfor_period_type = table.cell(datas + 1, head_data.index('周期类型') + 1).value if table.cell(datas + 1, head_data.index('周期类型') + 1).value is not None else ""
                                perfor_costs_attach = table.cell(datas + 1, head_data.index('成本归属') + 1).value if table.cell(datas + 1, head_data.index('成本归属') + 1).value is not None else ""
                                perfor_proof_ip = table.cell(datas + 1, head_data.index('个人绩效占比') + 1).value
                                perfor_proof_dp = table.cell(datas + 1, head_data.index('部门绩效占比') + 1).value
                                perfor_proof_cp = table.cell(datas + 1, head_data.index('公司绩效占比') + 1).value
                                perfor_score_ip = table.cell(datas + 1, head_data.index('个人绩效分数') + 1).value
                                perfor_score_dp = table.cell(datas + 1, head_data.index('部门绩效分数') + 1).value
                                perfor_score_cp = table.cell(datas + 1, head_data.index('公司绩效分数') + 1).value
                                perfor_ratio_ip = table.cell(datas + 1, head_data.index('个人绩效系数') + 1).value
                                perfor_ratio_dp = table.cell(datas + 1, head_data.index('部门绩效系数') + 1).value
                                perfor_ratio_cp = table.cell(datas + 1, head_data.index('公司绩效系数') + 1).value
                                perfor_result = table.cell(datas + 1, head_data.index('考核结果') + 1).value if table.cell(datas + 1, head_data.index('考核结果') + 1).value is not None else ""
                                WagePerformanceInto.objects.update_or_create(
                                    defaults={
                                        'perfor_emp_id': perfor_emp_id,
                                        'perfor_center': perfor_center,
                                        'perfor_periods': period_name,
                                        'perfor_period_type': perfor_period_type,
                                        'perfor_costs_attach': perfor_costs_attach,
                                        'perfor_proof_ip': perfor_proof_ip,
                                        'perfor_proof_dp': perfor_proof_dp,
                                        'perfor_proof_cp': perfor_proof_cp,
                                        'perfor_score_ip': perfor_score_ip,
                                        'perfor_score_dp': perfor_score_dp,
                                        'perfor_score_cp': perfor_score_cp,
                                        'perfor_ratio_ip': perfor_ratio_ip,
                                        'perfor_ratio_dp': perfor_ratio_dp,
                                        'perfor_ratio_cp': perfor_ratio_cp,
                                        'perfor_ratio_tp': perfor_proof_ip * perfor_ratio_ip + perfor_proof_dp * perfor_ratio_dp + perfor_proof_cp * perfor_ratio_cp if (perfor_proof_ip is not None and
                                                                                                                                                                         perfor_ratio_ip is not None and
                                                                                                                                                                         perfor_proof_dp is not None and
                                                                                                                                                                         perfor_ratio_dp is not None and
                                                                                                                                                                         perfor_proof_cp is not None and
                                                                                                                                                                         perfor_ratio_cp is not None) else None,
                                        'perfor_result': perfor_result,
                                        'perfor_creater': str(request.user),
                                        'perfor_create_time': firstdate.datetime.now()

                                    },
                                    perfor_emp_id=perfor_emp_id,
                                    perfor_periods=period_name,
                                )
                        else:
                            error_str += "第" + str(datas) + "行工号和姓名不匹配，请修正。"
                            continue
                    else:
                        error_str += "第" + str(datas) + "行有空字符串，请整行删除。"
                        continue
            # 如果错误
            if error_str == "":
                status = 'success'
                msg = '上传成功！'
            else:
                status = 'error'
                msg = error_str
            return JsonResponse(data={
                'status': status,
                'msg': str(msg)
            })
        else:
            status = 'error'
            msg = '列名错误，请使用下载的文档进行导入！'
            return JsonResponse(data={
                'status': status,
                'msg': str(msg)
            })

    upload_file.short_description = '文件上传'
    upload_file.type = 'success'
    upload_file.icon = 'el-icon-upload'
    upload_file.enable = True

    upload_file.layer = {
        'params': [{
            'type': 'file',
            'key': 'upload',
            'label': '文件'
        }]
    }

    def download(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        row1 = ['工号', '姓名', '中心别', '周期名称', '周期类型', '成本归属', '公司', '部门简称', '岗位名称', '个人绩效占比', '部门绩效占比', '公司绩效占比', '个人绩效分数', '部门绩效分数', '公司绩效分数', '个人绩效系数', '部门绩效系数', '公司绩效系数', '综合绩效系数', '考核结果']
        ws.append(row1)
        for datas in queryset:
            row2 = [
                int(datas.perfor_emp.emp_id),  # 工号
                str(datas.perfor_emp.emp_name),  # 姓名
                '',  # 中心别
                str(datas.perfor_periods),  # 周期名称
                str(datas.perfor_period_type),  # 周期类型
                str(datas.perfor_costs_attach),  # 成本归属
                str(datas.perfor_emp.emp_rank),  # 公司
                str(datas.perfor_emp.emp_dept.dept_short_name),  # 部门简称
                str(datas.perfor_emp.emp_posi),  # 岗位名称
                datas.perfor_proof_ip,  # 个人绩效指数占比
                datas.perfor_proof_dp,  # 部门绩效指数占比
                datas.perfor_proof_cp,  # 公司绩效指数占比
                datas.perfor_score_ip,  # 个人绩效分数
                datas.perfor_score_dp,  # 部门绩效分数
                datas.perfor_score_cp,  # 公司绩效分数
                datas.perfor_ratio_ip,  # 个人绩效系数
                datas.perfor_ratio_dp,  # 部门绩效系数
                datas.perfor_ratio_cp,  # 公司绩效系数
                datas.perfor_ratio_tp,  # 综合绩效系数
                str(datas.perfor_result),  # 考核结果
            ]
            ws.append(row2)
        wb.save('static/download/performance.xlsx')
        #     下载
        file = open('static/download/performance.xlsx', 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="performance.xlsx"'
        return response

    download.short_description = '下载数据'
    download.type = 'success'


# 固定工资引入表
class WageFixWageAdmin(AjaxAdmin):
    list_display = ['fix_emp', 'fix_emp_ids', 'dept', 'position', 'fix_begin_date', 'fix_work_pro', 'fix_wage_types',
                    'fix_skill_level', 'fix_manage_level', 'fix_wage_stand', 'fix_base_wage', 'fix_post_wage', 'fix_perfor_wage',
                    'fix_skill_wage', 'fix_manage_wage', 'fix_phone_subsidy', 'fix_change_reason', 'fix_resource']
    raw_id_fields = ['fix_emp']
    actions = ['upload_file', 'download_templates', 'download']
    search_fields = ['fix_emp__emp_id', 'fix_emp__emp_name']
    list_filter = ['fix_emp__emp_dept__dept_short_name', 'fix_emp__emp_posi']

    def fix_wage_types(self, obj):
        if obj.fix_emp.emp_pay_type == 1:
            return '计时制'
        else:
            return '责任制'

    fix_wage_types.short_description = '计薪方式'

    def get_list_display(self, request):
        myUser = MyUser.objects.get(user_id=request.user.id)
        fieldList = myUser.field.all().values_list('field_code')
        dataList = ['fix_emp', 'fix_emp_ids', 'dept', 'position', 'fix_begin_date', 'fix_work_pro', 'fix_wage_types', ]
        for data in fieldList:
            dataList.append(data[0])
        return dataList

    def get_fields(self, request, obj=None):
        myUser = MyUser.objects.get(user_id=request.user.id)
        fieldList = myUser.field.all().values_list('field_code')
        dataList = ['fix_emp', 'fix_begin_date', 'fix_work_pro', 'fix_wage_type', 'fix_change_reason', 'fix_resource', 'fix_creater', 'fix_create_time']
        for data in fieldList:
            dataList.append(data[0])
        return dataList

    def get_readonly_fields(self, request, obj=None):
        if request.user.is_superuser:
            return 'fix_creater', 'fix_create_time'
        else:
            return 'fix_creater', 'fix_create_time', 'fix_emp', 'fix_emp_ids', 'dept', 'position', 'fix_wage_type', 'fix_begin_date', 'fix_work_pro', 'fix_wage_types', 'fix_skill_level', 'fix_manage_level', 'fix_wage_stand', 'fix_base_wage', 'fix_post_wage', 'fix_perfor_wage', 'fix_skill_wage', 'fix_manage_wage', 'fix_phone_subsidy', 'fix_change_reason', 'fix_resource'

    def get_queryset(self, request):
        qs = super(WageFixWageAdmin, self).get_queryset(request)
        if not request.user.is_superuser:
            myUser = MyUser.objects.get(user_id=request.user.id)
            deptIdList = myUser.department.all().values_list('id')
            rankId = myUser.rank.all().values_list('id')
            if request.user.myuser.is_time is True:
                return qs.filter(fix_emp__emp_dept_id__in=deptIdList, fix_emp__emp_rank_id__in=rankId, fix_emp__emp_pay_type=1)
            else:
                return qs.filter(fix_emp__emp_dept_id__in=deptIdList, fix_emp__emp_rank_id__in=rankId)
        else:
            return qs

    # 根据登录账号判断显示的按钮
    def get_actions(self, request):
        actions = super(WageFixWageAdmin, self).get_actions(request)
        if not request.user.has_perm('wage.can_import_fix'):
            del actions['upload_file']
        return actions

    def fix_emp_ids(self, obj):
        return obj.fix_emp.emp_id

    fix_emp_ids.short_description = '工号'

    def dept(self, obj):
        return obj.fix_emp.emp_dept.dept_short_name

    dept.short_description = '部门'

    def position(self, obj):
        return obj.fix_emp.emp_posi

    position.short_description = '岗位'

    # def _filter_actions_by_permissions(self, request, actions):

    def upload_file(self, request, queryset):
        # 这里的upload 就是和params中配置的key一样
        upload = request.FILES['upload']
        now_time = str(firstdate.datetime.now().strftime("%Y-%m-%d_%H.%M.%S"))
        WageLog.objects.create(log_user=str(request.user), log_model='固定工资引入', log_path='static/fixwage/', log_file_name=now_time + str(request.user) + upload.name)
        # 将上传的文件保存到static中
        with open('static/fixwage/' + now_time + str(request.user) + upload.name, 'wb') as f:
            for line in upload.chunks():
                f.write(line)
        f.close()
        # try:
        workbook = openpyxl.load_workbook('static/fixwage/' + now_time + str(request.user) + upload.name)
        # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
        table = workbook.active
        rows = table.max_row
        cols = table.max_column
        # 获取标题栏数据
        head_data = []
        for datas in range(cols):
            head_data.append(table.cell(1, datas + 1).value)
        # 首先判断列名是否都正确，正确则进入循环，错误就返回报错
        if '工号' in head_data and '姓名' in head_data and '类型' in head_data and '起始日期' in head_data and '公司' in head_data and '工序' in head_data and '薪酬类型' in head_data and '技能等级' in head_data and '管理等级' in head_data and '标准工资' in head_data and '基本工资' in head_data and '岗位工资' in head_data and '绩效工资' in head_data and '技能津贴' in head_data and '管理津贴' in head_data and '话费补贴' in head_data and '变更原因' in head_data:
            # 创建一个字符串来手机所有错误数据
            error_str = ''
            # 列名都正确，开始遍历每行数据
            for datas in range(rows):
                # 从第二行开始写入数据库
                if datas > 0:
                    emp_id = table.cell(datas + 1, head_data.index('工号') + 1).value
                    # 先通过emp_id判断是否这行真有数据，没有的话则是空白格式，有的话就对比工号和姓名。
                    if type(emp_id) == int:
                        # 通过获取的工号去数据库中查找对应的姓名
                        sql_name = WageEmployee.objects.raw("select id,emp_name from wage_employee where emp_id = '" + str(emp_id) + "'")
                        emp_name = table.cell(datas + 1, head_data.index('姓名') + 1).value
                        # 如果此emp_id且有对应的emp_name则继续写入，没有的话则跳过当条循环，记录错误日志。继续下一条。
                        if len(sql_name) > 0 and str(sql_name[0]) == emp_name:
                            emp_ids = WageEmployee.objects.filter(emp_id=int(emp_id)).values('id')[0]['id']
                            change_type = table.cell(datas + 1, head_data.index('类型') + 1).value
                            begin_date = table.cell(datas + 1, head_data.index('起始日期') + 1).value
                            # 判断是否为日期格式（datetime），不是的话报错，是的话继续循环
                            if type(begin_date) == int:
                                error_str += "第" + str(datas) + "行日期格式错误，请以文本格式或日期格式，不要使用自定义格式。"
                                continue
                            else:
                                work_pro = table.cell(datas + 1, head_data.index('工序') + 1).value
                                wage_type = table.cell(datas + 1, head_data.index('薪酬类型') + 1).value
                                skill_level = table.cell(datas + 1, head_data.index('技能等级') + 1).value
                                manage_level = table.cell(datas + 1, head_data.index('管理等级') + 1).value
                                wage_stand = table.cell(datas + 1, head_data.index('标准工资') + 1).value
                                base_wage = table.cell(datas + 1, head_data.index('基本工资') + 1).value
                                post_wage = table.cell(datas + 1, head_data.index('岗位工资') + 1).value
                                perfor_wage = table.cell(datas + 1, head_data.index('绩效工资') + 1).value
                                # 如果是计时制的，就获取技能津贴和管理津贴
                                # 因技能等级、管理等级只有计时制有，系统添加自动识别功能，如责任制上传了技能等级或管理等级，报错
                                wages_type = WageEmployee.objects.raw("select id, emp_pay_type,emp_job_level_id from wage_employee where emp_id = %d" % emp_id)
                                if wages_type[0].emp_pay_type == 1:
                                    # 管理津贴
                                    if manage_level != "" and manage_level is not None:
                                        # 查询岗位名称
                                        posi_sql = "SELECT a.id, a.posi_name from wage_position as a INNER JOIN wage_employee as b on a.id = b.emp_posi_id where b.emp_id = '" + str(emp_id) + "'"
                                        emp_position = WagePosition.objects.raw(posi_sql)
                                        # 根据岗位和级别查询金额
                                        sql_manage = "select id,manage_allowance from wage_manage where instr('" + str(
                                            emp_position[0]) + "',manage_position) > 0 and manage_grade = '" + str(manage_level) + "'"
                                        manage_wage = WageManage.objects.raw(sql_manage)
                                        if len(manage_wage) > 0:
                                            manage_wages = manage_wage[0].manage_allowance
                                        else:
                                            manage_wages = 0
                                    else:
                                        manage_wages = 0
                                        manage_level = '无等级'
                                    # 技能津贴 根据基地和技能等级查找技能津贴
                                    if skill_level != "" and skill_level is not None:
                                        # 查询基地名
                                        # base_sql = "select a.id, a.rank_name from wage_rank as a INNER JOIN wage_employee as b on a.id = b.emp_rank_id where b.emp_id = '" + str(emp_id) + "'"
                                        # emp_rank = WageRank.objects.raw(base_sql)
                                        emp_rank = WageBase.objects.raw(
                                            "select a.id as id, a.dept_base as dept_base from wage_deptment as a left JOIN wage_employee as b on a.id = b.emp_dept_id where b.emp_id = '%s'" % str(
                                                emp_id))
                                        rank_id = WageBase.objects.raw("select id,base_name from wage_base where base_name = '" + str(emp_rank[0].dept_base) + "'")
                                        # 部门简称
                                        dept_short_name = WageDeptment.objects.raw(
                                            "select a.id as id,a.dept_short_name as short_name from wage_deptment  as a INNER JOIN wage_employee as b on a.id = b.emp_dept_id where b.emp_id = %d" % emp_id)
                                        # 职级
                                        job_ranks = wages_type[0].emp_job_level_id
                                        if len(rank_id) > 0:
                                            skill_wage = WageSkill.objects.raw(
                                                "select DISTINCT(a.id) as id,a.skill_grade,a.skill_allowance from wage_skill as a LEFT JOIN wage_skill_skill_dept as b on a.id = b.wageskill_id LEFT JOIN wage_skill_skill_local as c on a.id = c.wageskill_id LEFT JOIN wage_skill_skill_job_level as d on a.id = d.wageskill_id where b.wagedeptment_id = %s and c.wagebase_id = %s and d.wagejoblevel_id = %s and a.skill_grade ='%s'" % (
                                                    dept_short_name[0].id, rank_id[0].id, job_ranks, str(skill_level)))
                                            if len(skill_wage) > 0:
                                                skill_wages = skill_wage[0].skill_allowance
                                            else:
                                                skill_wages = 0
                                        else:
                                            skill_wages = 0
                                    else:
                                        skill_level = '无等级'
                                        skill_wages = 0
                                else:
                                    skill_level = '无等级'
                                    manage_level = '无等级'
                                    manage_wages = 0
                                    skill_wages = 0

                                phone_subsidy = table.cell(datas + 1, head_data.index('话费补贴') + 1).value
                                change_reason = table.cell(datas + 1, head_data.index('变更原因') + 1).value
                                defaultsList = {
                                    'fix_emp_id': emp_ids,
                                    'fix_change_type': change_type,
                                    'fix_begin_date': begin_date,
                                    'fix_work_pro': work_pro,
                                    'fix_wage_type': wage_type,
                                    'fix_change_reason': change_reason,
                                    'fix_resource': 'excel导入',
                                    'fix_creater': str(request.user),
                                    'fix_create_time': firstdate.datetime.now()
                                }
                                myUser = MyUser.objects.get(user_id=request.user.id)
                                fieldList = myUser.field.all().values_list('field_code')
                                for data in fieldList:
                                    if data[0] == 'fix_skill_level' and skill_level is not None:
                                        defaultsList.update({data[0]: skill_level})
                                        defaultsList.update({'fix_skill_wage': float(skill_wages)})
                                    # if data[0] == 'fix_manage_level' and manage_level is not None:
                                    if data[0] == 'fix_manage_level':
                                        defaultsList.update({data[0]: manage_level})
                                        defaultsList.update({'fix_manage_wage': float(manage_wages)})
                                    if data[0] == 'fix_wage_stand' and wage_stand is not None:
                                        defaultsList.update({data[0]: float(wage_stand)})
                                    if data[0] == 'fix_base_wage' and base_wage is not None:
                                        defaultsList.update({data[0]: float(base_wage)})
                                    if data[0] == 'fix_post_wage' and post_wage is not None:
                                        defaultsList.update({data[0]: float(post_wage)})
                                    if data[0] == 'fix_perfor_wage' and perfor_wage is not None:
                                        defaultsList.update({data[0]: float(perfor_wage)})
                                    if data[0] == 'fix_phone_subsidy' and phone_subsidy is not None:
                                        defaultsList.update({data[0]: float(phone_subsidy)})

                                WageFixWage.objects.update_or_create(
                                    defaults=defaultsList,
                                    fix_emp_id=emp_ids,
                                )
                                # WageFixLog.objects.create(log_creater=str(request.user),
                                #                           log_time=firstdate.datetime.now(),
                                #                           log_emp_id=emp_ids,
                                #                           log_change_type=change_type,
                                #                           log_begin_date=begin_date,
                                #                           log_work_pro=work_pro,
                                #                           log_wage_type=wage_type,
                                #                           log_skill_level=skill_level,
                                #                           log_manage_level=manage_level,
                                #                           log_wage_stand=wage_stand,
                                #                           log_base_wage=base_wage,
                                #                           log_post_wage=post_wage,
                                #                           log_perfor_wage=perfor_wage,
                                #                           log_skill_wage=skill_wages,
                                #                           log_manage_wage=manage_wages,
                                #                           log_phone_subsidy=phone_subsidy,
                                #                           log_change_reason=change_reason,
                                #                           log_resource='excel导入',
                                #                           )
                        else:
                            error_str += "第" + str(datas) + "行工号和姓名不匹配，请修正。"
                            continue
                    else:
                        error_str += "第" + str(datas) + "行有空字符串，请整行删除。"
                        continue
            # 如果错误
            if error_str == "":
                status = 'success'
                msg = '上传成功！'
            else:
                status = 'error'
                msg = error_str
            return JsonResponse(data={
                'status': status,
                'msg': str(msg)
            })
        # 列名错误，报错
        else:
            status = 'error'
            msg = '列名错误，请使用下载的文档进行导入！'
            return JsonResponse(data={
                'status': status,
                'msg': str(msg)
            })

    upload_file.short_description = '文件上传'
    upload_file.type = 'success'
    upload_file.icon = 'el-icon-upload'
    upload_file.enable = True

    upload_file.layer = {
        'params': [{
            'type': 'file',
            'key': 'upload',
            'label': '文件'
        }]
    }

    def download_templates(self, request, queryset):
        file = open('static/templates/fixwage.xlsx', 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="fixwage.xlsx"'
        return response

    download_templates.short_description = '下载模板'
    download_templates.type = 'warning'
    download_templates.icon = 'el-icon-paperclip'
    download_templates.acts_on_all = True

    def download(self, request, queryset):
        myUser = MyUser.objects.get(user_id=request.user.id)
        fieldList = myUser.field.all().values('field_code')
        fixList = []
        for data in fieldList:
            fixList.append(data['field_code'])
        wb = openpyxl.Workbook()
        ws = wb.active
        row1 = ['工号', '姓名', '类型', '起始日期', '公司', '部门简称', '岗位名称', '工序', '薪酬类型', '技能等级', '管理等级', '标准工资', '基本工资', '岗位工资', '绩效工资', '技能津贴', '管理津贴', '话费补贴', '变更原因']
        ws.append(row1)
        for datas in queryset:
            row2 = [
                int(datas.fix_emp.emp_id),  # 工号
                str(datas.fix_emp.emp_name),  # 姓名
                str(datas.fix_change_type) if str(datas.fix_change_type) != "None" else "",  # 类型
                str(datas.fix_begin_date),  # 起始日期
                str(datas.fix_emp.emp_rank),  # 公司
                str(datas.fix_emp.emp_dept.dept_short_name),  # 部门简称
                str(datas.fix_emp.emp_posi),  # 岗位名称
                str(datas.fix_work_pro) if str(datas.fix_work_pro) != "None" else "",  # 工序
                '计时制' if datas.fix_emp.emp_pay_type == 1 else '责任制',  # 薪酬类型
                datas.fix_skill_level if 'fix_skill_level' in fixList else "",  # 技能等级
                datas.fix_manage_level if 'fix_manage_level' in fixList else "",  # 管理等级
                datas.fix_wage_stand if 'fix_wage_stand' in fixList else "",  # 标准工资
                datas.fix_base_wage if 'fix_base_wage' in fixList else "",  # 基本工资
                datas.fix_post_wage if 'fix_post_wage' in fixList else "",  # 岗位工资
                datas.fix_perfor_wage if 'fix_perfor_wage' in fixList else "",  # 绩效工资
                datas.fix_skill_wage if 'fix_skill_wage' in fixList else "",  # 技能津贴
                datas.fix_manage_wage if 'fix_manage_wage' in fixList else "",  # 管理津贴
                datas.fix_phone_subsidy if 'fix_phone_subsidy' in fixList else "",  # 话费补贴
                datas.fix_change_reason,  # 变更原因
            ]
            ws.append(row2)
        wb.save('static/download/fixwage.xlsx')
        #     下载
        file = open('static/download/fixwage.xlsx', 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="fixwage.xlsx"'
        return response

    download.short_description = '下载数据'
    download.type = 'success'


# 月度工资汇总表
class WageMonthAdmin(AjaxAdmin):
    list_display = ['mon_emp', 'dept', 'position', 'mon_periods', 'mon_wage_way', 'mon_standard_wage', 'mon_base_wage',
                    'mon_post_wage', 'mon_perfor_wage', 'mon_manage_wage', 'mon_seniority_wage', 'mon_post_subsidy',
                    'mon_skill_wage', 'mon_perfor_merit', 'mon_reward_punish', 'mon_other_deduc', 'mon_si_af_deduc',
                    'mon_reward', 'mon_phyexam', 'mon_water_elect', 'mon_rent', 'mon_quart_bonus', 'mon_annual_bonus',
                    'mon_sale_com', 'mon_temp_wage', 'mon_laber_cost', 'mon_income_tax', 'mon_non_compet',
                    'mon_severance']
    raw_id_fields = ['mon_emp']
    search_fields = ['mon_emp__emp_id', 'mon_emp__emp_name']
    list_filter = ['mon_periods', 'mon_emp__emp_dept__dept_short_name', 'mon_emp__emp_posi']
    actions = ['layer_input', 'download']
    list_per_page = 10

    def get_actions(self, request):
        actions = super(WageMonthAdmin, self).get_actions(request)
        if not request.user.has_perm('wage.can_export_month'):
            del actions['download']
        return actions

    def get_queryset(self, request):
        qs = super(WageMonthAdmin, self).get_queryset(request)
        if not request.user.is_superuser:
            myUser = MyUser.objects.get(user_id=request.user.id)
            deptIdList = myUser.department.all().values_list('id')
            rankId = myUser.rank.all().values_list('id')
            return qs.filter(mon_emp__emp_dept_id__in=deptIdList, mon_emp__emp_rank_id__in=rankId)
        else:
            return qs

    def dept(self, obj):
        return obj.mon_emp.emp_dept.dept_short_name

    dept.short_description = '部门'

    def position(self, obj):
        return obj.mon_emp.emp_posi

    position.short_description = '岗位'

    def layer_input(self, request, queryset):
        # 这里的queryset 会有数据过滤，只包含选中的数据
        post = request.POST
        monthcalc.calc(post.get('type'))
        if not post.get('type'):
            return JsonResponse(data={
                'status': 'error',
                'msg': '请先选中数据！'
            })
        else:
            return JsonResponse(data={
                'status': 'success',
                'msg': '处理成功！'
            })

    layer_input.short_description = '生成报表'
    layer_input.type = 'success'
    layer_input.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    layer_input.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        'title': '请选择',
        # 提示信息
        'tips': '请选择需要生成薪酬的周期',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [{
            'type': 'month',
            'key': 'type',
            'label': '周期',
            'width': '200px',
            # size对应elementui的size，取值为：medium / small / mini
            'size': 'small',
            # value字段可以指定默认值
            # 'value': '0',
            # 'options': list(WagePeriod.objects.annotate(key=F('id'), label=F('period_name')).values('key', 'label'))
        }, ]
    }

    def download(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        row1 = ['工号', '姓名', '集团入职日期', '入职日期', '离职结薪日期', '离职办理日期', '公司', '部门', '岗位', '计薪周期', '计薪方式', '标准工资', '基本工资', '岗位工资', '绩效工资', '管理津贴', '年资津贴', '岗位补贴',
                '技能津贴', '绩效系数', '奖惩', '其他补扣款', '社保公积金补扣款', '奖金/津贴', '体检费报销', '代收水电费', '代收房租', '季度奖金', '年度奖金', '销售提成',
                '临时补贴', '劳务费', '个人所得税', '竞业限制补偿金', '离职补偿金'
                ]
        ws.append(row1)
        for datas in queryset:
            row2 = [
                str(datas.mon_emp.emp_id) if str(datas.mon_emp.emp_id) != "None" else "",  # 工号
                str(datas.mon_emp.emp_name) if str(datas.mon_emp.emp_name) != "None" else "",  # 姓名
                str(datas.mon_emp.emp_entry_date) if str(datas.mon_emp.emp_entry_date) != "None" else "",  # 集团入职日期
                str(datas.mon_emp.emp_in_date) if str(datas.mon_emp.emp_in_date) != "None" else "",  # 入职日期
                str(datas.mon_emp.emp_leave_date) if str(datas.mon_emp.emp_leave_date) != "None" else "",  # 离职结薪日期
                str(datas.mon_emp.emp_leave_bl_date) if str(datas.mon_emp.emp_leave_bl_date) != "None" else "",  # 离职办理日期
                str(datas.mon_emp.emp_dept.dept_base) if str(datas.mon_emp.emp_dept.dept_base) != "None" else "",  # 公司
                str(datas.mon_emp.emp_dept.dept_name) if str(datas.mon_emp.emp_dept.dept_name) != "None" else "",  # 部门
                str(datas.mon_emp.emp_posi.posi_name) if str(datas.mon_emp.emp_posi.posi_name) != "None" else "",  # 岗位
                str(datas.mon_periods) if str(datas.mon_periods) != "None" else "",  # 计薪周期
                str(datas.mon_wage_way) if str(datas.mon_wage_way) != "None" else "",  # 计薪方式
                str(datas.mon_standard_wage) if str(datas.mon_standard_wage) != "None" else 0,  # 标准工资
                str(datas.mon_base_wage) if str(datas.mon_base_wage) != "None" else 0,  # 基本工资
                str(datas.mon_post_wage) if str(datas.mon_post_wage) != "None" else 0,  # 岗位工资
                str(datas.mon_perfor_wage) if str(datas.mon_perfor_wage) != "None" else 0,  # 绩效工资
                str(datas.mon_manage_wage) if str(datas.mon_manage_wage) != "None" else 0,  # 管理津贴
                str(datas.mon_seniority_wage) if str(datas.mon_seniority_wage) != "None" else 0,  # 年资津贴
                str(datas.mon_post_subsidy) if str(datas.mon_post_subsidy) != "None" else 0,  # 岗位补贴
                str(datas.mon_skill_wage) if str(datas.mon_skill_wage) != "None" else 0,  # 技能津贴
                str(datas.mon_perfor_merit) if str(datas.mon_perfor_merit) != "None" else "",  # 绩效系数
                str(datas.mon_reward_punish) if str(datas.mon_reward_punish) != "None" else 0,  # 奖惩
                str(datas.mon_other_deduc) if str(datas.mon_other_deduc) != "None" else 0,  # 其他补扣款
                str(datas.mon_si_af_deduc) if str(datas.mon_si_af_deduc) != "None" else 0,  # 社保公积金补扣款
                str(datas.mon_reward) if str(datas.mon_reward) != "None" else 0,  # 奖金/津贴
                str(datas.mon_phyexam) if str(datas.mon_phyexam) != "None" else 0,  # 体检费报销
                str(datas.mon_water_elect) if str(datas.mon_water_elect) != "None" else 0,  # 代收水电费
                str(datas.mon_rent) if str(datas.mon_rent) != "None" else 0,  # 代收房租
                str(datas.mon_quart_bonus) if str(datas.mon_quart_bonus) != "None" else 0,  # 季度奖金
                str(datas.mon_annual_bonus) if str(datas.mon_annual_bonus) != "None" else 0,  # 年度奖金
                str(datas.mon_sale_com) if str(datas.mon_sale_com) != "None" else 0,  # 销售提成
                str(datas.mon_temp_wage) if str(datas.mon_temp_wage) != "None" else 0,  # 临时补贴
                str(datas.mon_laber_cost) if str(datas.mon_laber_cost) != "None" else 0,  # 劳务费
                str(datas.mon_income_tax) if str(datas.mon_income_tax) != "None" else 0,  # 个人所得税
                str(datas.mon_non_compet) if str(datas.mon_non_compet) != "None" else 0,  # 竞业限制补偿金
                str(datas.mon_severance) if str(datas.mon_severance) != "None" else 0,  # 离职补偿金
            ]
            ws.append(row2)
        wb.save('static/data.xlsx')
        #     下载
        file = open('static/data.xlsx', 'rb')
        response = FileResponse(file)
        response['Content-Type'] = 'application/octet-stream'
        response['Content-Disposition'] = 'attachment;filename="record.xlsx"'
        return response

        # 显示的文本，与django admin一致

    download.short_description = '下载数据'
    # icon，参考element-ui icon与https://fontawesome.com
    # custom_button.icon = 'fas fa-audio-description'
    # 指定element-ui的按钮类型，参考https://element.eleme.cn/#/zh-CN/component/button
    download.type = 'success'
    download.confirm = '是否确定下载此报表？'

    # 给按钮追加自定义的颜色
    # custom_button.style = 'color:black;'


# 基地表
class WageBaseAdmin(AjaxAdmin):
    list_display = ['base_name']
    search_fields = ['base_name']


# 技能津贴表
class WageSkillAdmin(AjaxAdmin):
    list_display = ['skill_grade', 'skill_allowance', 'skill_locals', 'skill_depts', 'skill_job_levels']
    list_filter = ['skill_grade', 'skill_local__base_name']
    list_display_links = ['skill_grade', 'skill_allowance', 'skill_locals', 'skill_depts', 'skill_job_levels']
    search_fields = ['skill_locals', 'skill_local__base_name']

    def skill_locals(self, obj):
        tag_list = []
        tags = obj.skill_local.all()
        if tags:
            for tag in tags:
                tag_list.append(tag.base_name)
            return ','.join(tag_list)
        else:
            return format_html(
                '<span style="color:red;">文章{}无标签</span>',
                obj.id, )

    skill_locals.short_description = '基地'

    def ifNull(self, data):
        if data != "" and data is not None:
            return data
        else:
            return "..."

    def skill_depts(self, obj):
        tag_list = []
        tags = obj.skill_dept.all()
        if tags:
            for tag in tags:
                tag_list.append(self.ifNull(tag.dept_name) + '||' + self.ifNull(tag.dept_short_name))
            return ','.join(tag_list)
        else:
            return format_html(
                '<span style="color:red;">文章{}无标签</span>',
                obj.id, )

    skill_depts.short_description = '部门'

    def skill_job_levels(self, obj):
        tag_list = []
        tags = obj.skill_job_level.all()
        if tags:
            for tag in tags:
                tag_list.append(tag.level_name)
            return ','.join(tag_list)
        else:
            return format_html(
                '<span style="color:red;">文章{}无标签</span>',
                obj.id, )

    skill_job_levels.short_description = '职级'


# 管理津贴表
class WageManageAdmin(AjaxAdmin):
    list_display = ['manage_position', 'manage_grade', 'manage_allowance']
    list_filter = ['manage_position', 'manage_grade']


# 离职人员薪资
class WageLeaveAdmin(AjaxAdmin):
    list_display = ['leave_wage_period', 'leave_tel', 'leave_emp_code', 'leave_emp_id', 'leave_emp_name', 'leave_emp_dept', 'leave_emp_post', 'leave_wage_date', 'leave_base_wage',
                    'leave_position_wage', 'leave_perfor_wage', 'leave_phone', 'leave_skill', 'leave_manage', 'leave_year', 'leave_fix', 'leave_eat',
                    'leave_s_base_wage', 'leave_s_position_wage', 'leave_perfor_ratio', 'leave_perfor', 'leave_subsidy', 'leave_over', 'leave_night', 'leave_rp', 'leave_reward', 'leave_temp',
                    'leave_test',
                    'leave_attend', 'leave_holiday', 'leave_pfund', 'leave_other', 'leave_compet', 'leave_total_wage', 'leave_person_insur', 'leave_comp_insur', 'leave_income_tax', 'leave_we',
                    'leave_house',
                    'leave_otherplus', 'leave_othersub', 'leave_disp', 'leave_special', 'leave_actual_wage', 'leave_remark']
    search_fields = ['leave_emp_code', 'leave_emp_id', 'leave_emp_name']
    list_filter = ['leave_wage_period', 'leave_emp_dept', 'leave_emp_post']
    list_per_page = 10
    actions = ['upload_file', ]

    def get_actions(self, request):
        actions = super(WageLeaveAdmin, self).get_actions(request)
        if not request.user.has_perm('wage.can_import_month'):
            del actions['upload_file']
        return actions

    def upload_file(self, request, queryset):
        # 这里的upload 就是和params中配置的key一样
        upload = request.FILES['upload']
        # 将上传的文件保存到static中
        with open('static/leavewage/' + str(firstdate.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")) + str(request.user) + upload.name, 'wb') as f:
            for line in upload.chunks():
                f.write(line)
        f.close()
        workbook = openpyxl.load_workbook('static/leavewage/' + str(firstdate.datetime.now().strftime("%Y-%m-%d_%H.%M.%S")) + str(request.user) + upload.name)
        # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
        table = workbook.active
        rows = table.max_row
        cols = table.max_column
        # 获取标题栏数据
        head_data = []
        for datas in range(cols):
            head_data.append(table.cell(1, datas + 1).value)
        # 创建一个字符串来手机所有错误数据
        error_str = ''
        for datas in range(rows):
            if datas > 0:
                leave_emp_code = table.cell(datas + 1, head_data.index('工号') + 1).value
                if type(leave_emp_code) == int:
                    leave_wage_period = table.cell(datas + 1, head_data.index('周期') + 1).value
                    # 判断是否为日期格式（datetime），不是的话报错，是的话继续循环
                    if type(leave_wage_period) == int:
                        error_str += "第" + str(datas) + "行日期格式错误，请以文本格式，不要使用自定义格式。"
                        continue
                    else:
                        leave_tel = table.cell(datas + 1, head_data.index('手机号码') + 1).value
                        leave_emp_id = table.cell(datas + 1, head_data.index('身份证号') + 1).value
                        leave_emp_name = table.cell(datas + 1, head_data.index('员工姓名') + 1).value
                        leave_emp_dept = table.cell(datas + 1, head_data.index('部门简称') + 1).value
                        leave_emp_post = table.cell(datas + 1, head_data.index('岗位名称') + 1).value
                        leave_wage_date = table.cell(datas + 1, head_data.index('离职结薪日期') + 1).value
                        leave_base_wage = table.cell(datas + 1, head_data.index('基本工资') + 1).value
                        leave_position_wage = table.cell(datas + 1, head_data.index('岗位工资') + 1).value
                        leave_perfor_wage = table.cell(datas + 1, head_data.index('绩效工资') + 1).value
                        leave_phone = table.cell(datas + 1, head_data.index('话费补贴') + 1).value
                        leave_skill = table.cell(datas + 1, head_data.index('技能津贴') + 1).value
                        leave_manage = table.cell(datas + 1, head_data.index('管理津贴') + 1).value
                        leave_year = table.cell(datas + 1, head_data.index('年资津贴') + 1).value
                        leave_fix = table.cell(datas + 1, head_data.index('固定补贴') + 1).value
                        leave_eat = table.cell(datas + 1, head_data.index('就餐补贴') + 1).value
                        leave_s_base_wage = table.cell(datas + 1, head_data.index('应发基本工资') + 1).value
                        leave_s_position_wage = table.cell(datas + 1, head_data.index('应发岗位工资') + 1).value
                        leave_perfor_ratio = table.cell(datas + 1, head_data.index('绩效综合系数') + 1).value
                        leave_perfor = table.cell(datas + 1, head_data.index('应发绩效工资') + 1).value
                        leave_subsidy = table.cell(datas + 1, head_data.index('应发补贴合计') + 1).value
                        leave_over = table.cell(datas + 1, head_data.index('加班1.5') + 1).value + table.cell(datas + 1, head_data.index('加班2.0') + 1).value + table.cell(datas + 1, head_data.index(
                            '加班3.0') + 1).value
                        leave_night = table.cell(datas + 1, head_data.index('夜班津贴') + 1).value
                        leave_rp = table.cell(datas + 1, head_data.index('奖/惩') + 1).value
                        leave_reward = table.cell(datas + 1, head_data.index('奖金/津贴') + 1).value
                        leave_temp = table.cell(datas + 1, head_data.index('临时补贴') + 1).value
                        leave_test = table.cell(datas + 1, head_data.index('体检报销费用') + 1).value
                        leave_attend = table.cell(datas + 1, head_data.index('考勤异常扣款') + 1).value
                        leave_holiday = table.cell(datas + 1, head_data.index('事假扣款') + 1).value
                        leave_pfund = table.cell(datas + 1, head_data.index('社保公积金补扣') + 1).value
                        leave_other = table.cell(datas + 1, head_data.index('其他补扣款') + 1).value
                        leave_compet = table.cell(datas + 1, head_data.index('竞业补偿金') + 1).value
                        leave_total_wage = table.cell(datas + 1, head_data.index('应发合计') + 1).value
                        leave_person_insur = table.cell(datas + 1, head_data.index('社保个人合计') + 1).value
                        leave_comp_insur = table.cell(datas + 1, head_data.index('社保单位合计') + 1).value
                        leave_income_tax = table.cell(datas + 1, head_data.index('代扣个税') + 1).value
                        leave_we = table.cell(datas + 1, head_data.index('水电费') + 1).value
                        leave_house = table.cell(datas + 1, head_data.index('房租费') + 1).value
                        leave_otherplus = table.cell(datas + 1, head_data.index('其他+') + 1).value
                        leave_othersub = table.cell(datas + 1, head_data.index('其他-') + 1).value
                        leave_disp = table.cell(datas + 1, head_data.index('一次性补偿金') + 1).value
                        leave_special = table.cell(datas + 1, head_data.index('附加专项扣除合计') + 1).value
                        leave_actual_wage = table.cell(datas + 1, head_data.index('实发工资') + 1).value
                        leave_remark = table.cell(datas + 1, head_data.index('薪资备注') + 1).value

                        WageLeave.objects.update_or_create(
                            defaults={
                                'leave_wage_period': leave_wage_period,
                                'leave_tel': leave_tel,
                                'leave_emp_code': leave_emp_code,
                                'leave_emp_id': leave_emp_id,
                                'leave_emp_name': leave_emp_name,
                                'leave_emp_dept': leave_emp_dept,
                                'leave_emp_post': leave_emp_post,
                                'leave_wage_date': leave_wage_date,
                                'leave_base_wage': leave_base_wage,
                                'leave_position_wage': leave_position_wage,
                                'leave_perfor_wage': leave_perfor_wage,
                                'leave_phone': leave_phone,
                                'leave_skill': leave_skill,
                                'leave_manage': leave_manage,
                                'leave_year': leave_year,
                                'leave_fix': leave_fix,
                                'leave_eat': leave_eat,
                                'leave_s_base_wage': leave_s_base_wage,
                                'leave_s_position_wage': leave_s_position_wage,
                                'leave_perfor_ratio': leave_perfor_ratio,
                                'leave_perfor': leave_perfor,
                                'leave_subsidy': leave_subsidy,
                                'leave_over': leave_over,
                                'leave_night': leave_night,
                                'leave_rp': leave_rp,
                                'leave_reward': leave_reward,
                                'leave_temp': leave_temp,
                                'leave_test': leave_test,
                                'leave_attend': leave_attend,
                                'leave_holiday': leave_holiday,
                                'leave_pfund': leave_pfund,
                                'leave_other': leave_other,
                                'leave_compet': leave_compet,
                                'leave_total_wage': leave_total_wage,
                                'leave_person_insur': leave_person_insur,
                                'leave_comp_insur': leave_comp_insur,
                                'leave_income_tax': leave_income_tax,
                                'leave_we': leave_we,
                                'leave_house': leave_house,
                                'leave_otherplus': leave_otherplus,
                                'leave_othersub': leave_othersub,
                                'leave_disp': leave_disp,
                                'leave_special': leave_special,
                                'leave_actual_wage': leave_actual_wage,
                                'leave_remark': leave_remark
                            },
                            leave_emp_code=leave_emp_code,
                            leave_wage_period=leave_wage_period
                        )

                else:
                    error_str += "第" + str(datas) + "行有空字符串，请整行删除。"
                    continue
        # 如果错误
        if error_str == "":
            status = 'success'
            msg = '上传成功！'
        else:
            status = 'error'
            msg = error_str
        return JsonResponse(data={
            'status': status,
            'msg': str(msg)
        })

    upload_file.short_description = '文件上传'
    upload_file.type = 'success'
    upload_file.icon = 'el-icon-upload'
    upload_file.enable = True

    upload_file.layer = {
        'params': [{
            'type': 'file',
            'key': 'upload',
            'label': '文件'
        }]
    }


class LeaveReasonAdmin(AjaxAdmin):
    list_display = ['leave_reason_name']
    search_fields = ['leave_reason_name']


class LeaveAccountAdmin(AjaxAdmin):
    list_display = ['leave_emp', 'leave_dates', 'leave_leading_man', 'leave_prepare_leave_date', 'leave_reason', 'leave_content', 'leave_is_success', 'leave_is_other',
                    'leave_remarks', 'leave_suggests']
    search_fields = ['leave_emp__emp_id', 'leave_emp__emp_name']
    list_filter = ['leave_dates', 'leave_is_success', 'leave_reason', 'leave_emp__emp_dept', 'leave_emp__emp_posi', 'leave_emp__emp_dept__dept_base']
    raw_id_fields = ['leave_emp']
    list_per_page = 10

    def get_readonly_fields(self, request, obj=None):
        return 'leave_creater', 'leave_create_time'


class EmployeeInline(admin.StackedInline):
    model = MyUser
    can_delete = False
    verbose_name_plural = 'myuser'


# Define a new User admin
class UserAdmin(BaseUserAdmin):
    inlines = (EmployeeInline,)


class RecomdStandardsAdmin(AjaxAdmin):
    list_display = ['standards_effect_date', 'standards_invalid_date', 'job_levels', 'standards_zg_degree', 'standards_first_month', 'standards_second_month', 'standards_third_month',
                    'standards_four_month', 'standards_five_month', 'standards_six_month', 'standards_first_money', 'standards_second_money', 'standards_four_money', 'standards_five_money',
                    'standards_six_money', 'standards_first_money_b', 'standards_second_money_b', 'standards_four_money_b', 'standards_five_money_b', 'standards_six_money_b']
    list_display_links = ['standards_effect_date', 'standards_invalid_date', 'job_levels', 'standards_zg_degree', 'standards_first_month', 'standards_second_month', 'standards_third_month',
                          'standards_four_month', 'standards_five_month', 'standards_six_month', 'standards_first_money', 'standards_second_money', 'standards_four_money', 'standards_five_money',
                          'standards_six_money', 'standards_first_money_b', 'standards_second_money_b', 'standards_four_money_b', 'standards_five_money_b', 'standards_six_money_b']
    list_filter = ['standards_job_levels']
    actions = ['upload_file']

    def job_levels(self, obj):
        return [datas.level_name for datas in obj.standards_job_levels.all()]

    job_levels.short_description = '职级'

    def upload_file(self, request, queryset):
        # 这里的upload 就是和params中配置的key一样
        upload = request.FILES['upload']
        print(upload)
        pass

    upload_file.short_description = '文件上传'
    upload_file.type = 'success'
    upload_file.icon = 'el-icon-upload'
    upload_file.enable = True

    upload_file.layer = {
        'params': [{
            'type': 'file',
            'key': 'upload',
            'label': '文件'
        }]
    }


class RecomdCandidatesAdmin(AjaxAdmin):
    list_display = ['Candidates_emp', 'Candidates_emp_b', 'Candidates_native', 'Candidates_peer_exper', 'Candidates_peer', 'Candidates_status']
    list_display_links = ['Candidates_emp', 'Candidates_emp_b', 'Candidates_native', 'Candidates_peer_exper', 'Candidates_peer', 'Candidates_status']


class WageLogAdmin(AjaxAdmin):
    list_display = ['log_user', 'log_model', 'log_path', 'log_file_name', 'log_create_time']
    list_filter = ['log_user', 'log_model', 'log_create_time']


admin.site.site_header = '润阳人事管理平台'
admin.site.site_title = '润阳人事管理平台'
admin.site.index_title = '欢迎使用润阳人事管理平台'

admin.site.register(WageDeptment, WageDeptmentAdmin)
admin.site.register(WagePosition, WagePositionAdmin)
admin.site.register(WageEmployee, WageEmployeeAdmin)
admin.site.register(WageExpense, WageExpenseAdmin)
admin.site.register(WageExpenseInto, WageExpenseIntoAdmin)
admin.site.register(WagePerformanceInto, WagePerformanceIntoAdmin)
admin.site.register(WageFixWage, WageFixWageAdmin)
admin.site.register(WageMonth, WageMonthAdmin)
admin.site.register(WagePeriod, WagePeriodAdmin)
admin.site.register(WageBase, WageBaseAdmin)
admin.site.register(WageSkill, WageSkillAdmin)
admin.site.register(WageManage, WageManageAdmin)
admin.site.register(WageLeave, WageLeaveAdmin)
admin.site.register(LeaveReason, LeaveReasonAdmin)
admin.site.register(LeaveAccount, LeaveAccountAdmin)
admin.site.unregister(User)
admin.site.register(User, UserAdmin)
admin.site.register(RecomdStandards, RecomdStandardsAdmin)
admin.site.register(RecomdCandidates, RecomdCandidatesAdmin)
admin.site.register(WageLog, WageLogAdmin)
# admin.site.unregister(Group)
# admin.site.unregister(User)
